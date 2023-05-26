from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium import webdriver
from time import sleep as pause
from random import randint
from selenium.webdriver.common.action_chains import ActionChains
import re
import urllib.parse
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def get_links(text):

    with webdriver.Chrome() as driver:
        driver.maximize_window()
        url = 'https://hh.ru/?hhtmFrom=vacancy_search_list'
        driver.get(url)
        pause(randint(4,5))
        search = driver.find_element(By.ID, 'a11y-search-input').send_keys(text)

        button = driver.find_element(By.CLASS_NAME, 'supernova-search-submit-text')
        button.click()
        pause(randint(4,5))
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            pause(randint(4,5))

            pause(randint(1,3))
            soup = BeautifulSoup(driver.page_source, 'lxml')
            last_page = str(soup.find("div", class_='pager'))
            #last_page = f'https://hh.ru/search/vacancy?text={text}'

            pagination = [int(s) for s in re.findall(r"<span>(\d+)</span>", last_page)]

            encoded_string = text
            decoded_string = urllib.parse.quote(encoded_string)

            total_list_page = []
            for page in range(max(pagination)):
                x = f'https://hh.ru/search/vacancy?text={decoded_string}&page={page}'
                total_list_page.append(x)

            pause(2)
            return total_list_page

        except Exception:

            encoded_string = text
            decoded_string = urllib.parse.quote(encoded_string)
            current_url = f'https://hh.ru/search/vacancy?text={decoded_string}'
            #current_url = driver.current_url
            print(current_url)
            driver.get(current_url)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            pause(5)
            soup1 = BeautifulSoup(driver.page_source, 'lxml')
            total_product = []
            for title in soup1.find_all('div', class_="vacancy-serp-item-body"):
                product = {}
                try:
                    product['вакансия'] = title.find('a', class_="serp-item__title").text
                    #print(product['вакансия'])
                except:
                    continue
                product['работодатель'] = title.find('div', class_="bloko-text").text
                #print(product['работодатель'])
                product['ссылка'] = title.find('a', class_="serp-item__title").get('href')
                #print(product['ссылка'])
                for x in title.find_all('div', class_='vacancy-serp-item__info'):
                    div_elements = x.find_all('div', class_='bloko-text')
                    product['город'] = div_elements[-1].text


                total_product.append(product)
            df = pd.DataFrame(data=total_product, columns=['вакансия', 'работодатель', 'ссылка', 'город'])
            wb = Workbook()
            ws = wb.active

            for col, column_name in enumerate(df.columns):
                ws.cell(row=1, column=col+1, value=column_name)


            for r, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
                for c, value in enumerate(row):
                    ws.cell(row=r+2, column=c+1, value=value)


            for r in range(2, len(df)+2):
                link = df['ссылка'][r-2]
                ws.cell(row=r, column=3).hyperlink = link
                ws.cell(row=r, column=3).value = '=HYPERLINK("{}")'.format(link)

            #чтобы вводить название вручную поменять нужно закомментить следующую строку и раскомментить после нее
            filename = f'{text}.xlsx'
            #filename = input('введите название файла :') + '.xlsx'
            wb.save(filename=filename)
            print('найдена только одна страница с вакансиями :' , current_url)

def get_resume(link):
    total_result = []

    with webdriver.Chrome() as driver:
        driver.maximize_window()
        count = 1
        for page in link:

            url = page
            driver.get(url)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            pause(5)
            soup1 = BeautifulSoup(driver.page_source, 'lxml')


            for title in soup1.find_all('div', class_="vacancy-serp-item-body"):
                product = {}
                try:
                    product['вакансия'] = title.find('a', class_="serp-item__title").text

                except:
                    continue
                product['работодатель'] = title.find('div', class_="bloko-text").text

                product['ссылка'] = title.find('a', class_="serp-item__title").get('href')

                for x in title.find_all('div', class_='vacancy-serp-item__info'):
                    div_elements = x.find_all('div', class_='bloko-text')
                    product['город'] = div_elements[-1].text
                print('обработана карточка :', count)

                count += 1

                total_result.append(product)

    return total_result

if __name__== "__main__":
    input_work = input('Вакансия : ')
    result = get_links(input_work)
    if result:
        print('Найдено страниц с необходимой вакансией :', len(result), result)
        total_product = get_resume(result)
        df = pd.DataFrame(data=total_product, columns=['вакансия', 'работодатель', 'ссылка', 'город'])

        wb = Workbook()
        ws = wb.active

        for col, column_name in enumerate(df.columns):
            ws.cell(row=1, column=col+1, value=column_name)


        for r, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
            for c, value in enumerate(row):
                ws.cell(row=r+2, column=c+1, value=value)


        for r in range(2, len(df)+2):
            link = df['ссылка'][r-2]
            ws.cell(row=r, column=3).hyperlink = link
            ws.cell(row=r, column=3).value = '=HYPERLINK("{}")'.format(link)

        #чтобы вводить название вручную поменять нужно закомментить следующую строку и раскомментить после нее
        filename = f'{input_work}.xlsx'
        #filename = input('введите название файла :') + '.xlsx'
        wb.save(filename=filename)


    else:
        print('вакансий найдено не много, уточните поиск')

print("готово  "*20)